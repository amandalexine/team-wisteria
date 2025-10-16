# --------------------------------SAVE FUNCTION-----------------------------------------
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import io

import os

# Saves the graphs to the users excel file
def save_graphs_to_excel(excel_file, graphs_dict):
    temp_files = []
    
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "Stats Graphs" not in wb.sheetnames:
        ws = wb.create_sheet(title="Stats Graphs")
    else:
        ws = wb["Stats Graphs"]
        ws.delete_rows(1, ws.max_row)

    img_width = 600
    img_height = 400
    row_height = int(img_height / 1.5)
    current_row = 1

    for section, fig_list in graphs_dict.items():
        if not fig_list:
            continue
            
        cell = ws.cell(row=current_row, column=1, value=section)
        cell.font = Font(bold=True, size=12)  
        current_row += 1

        for i, fig in enumerate(fig_list, 1):
            if fig is None:  
                continue
            
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
            buf.seek(0)

            img = Image(buf)
            
            img.width = img_width
            img.height = img_height
            ws.add_image(img, f'A{current_row}')
            
            ws.row_dimensions[current_row].height = row_height + 5
            current_row += 5

        current_row += 2

    ws.column_dimensions['A'].width = img_width / 7
    wb.save(excel_file)

# Saves the ml graphs to the user's excel file
def save_ml_graphs_to_excel(excel_file, graphs_dict):
    temp_files = []
    
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "Signal Graphs" not in wb.sheetnames:
        ws = wb.create_sheet(title="Signal Graphs")
    else:
        ws = wb["Signal Graphs"]
        ws.delete_rows(1, ws.max_row)

    img_width = 1300
    img_height = 200
    row_height = int(img_height / 1.5)
    current_row = 1

    for category, data_types in graphs_dict.items():
        for data_type, fig in data_types.items():
            if fig is None:  
                continue

            title = category + "_" + data_type
            
            cell = ws.cell(row=current_row, column=1, value=title)
            cell.font = Font(bold=True, size=12)  
            current_row += 1

            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
            buf.seek(0)

            img = Image(buf)
            img.width = img_width
            img.height = img_height
            ws.add_image(img, f'A{current_row}')
            
            ws.row_dimensions[current_row].height = row_height + 5
            current_row += 5
        current_row += 2

    ws.column_dimensions['A'].width = img_width / 7
    wb.save(excel_file)

    for temp_file in temp_files:
        try:
            if os.path.exists(temp_file):
                os.remove(temp_file)
        except Exception as e:
            print(f"Error deleting temp file {temp_file}: {str(e)}")

# Saves the ml results to the excel file
def save_ml_results_to_excel(excel_file, ml_prediction, ml_features):
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "ML Results" not in wb.sheetnames:
        ws = wb.create_sheet(title="ML Results")
    else:
        ws = wb["ML Results"]
        ws.delete_rows(1, ws.max_row)


    classification = ml_prediction['ecg'].get('classification', 'N/A')
    confidence = ml_prediction['ecg'].get('confidence', 'N/A')

    if isinstance(confidence, (int, float)) and classification == 0:
        confidence = 1 - confidence

    
    
    sheet = wb['ML Results']
    sheet.cell(row=1, column=1).value = 'ML Classification:'
    sheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=1, column=2).value = classification

    sheet.cell(row=2, column=1).value = 'ML Confidence:'
    sheet.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=2, column=2).value = confidence

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20

    curr_row = 4
    if ml_prediction['ecg'] is not None:

        for signal_type, categories in ml_features.items():
            sheet.cell(row=curr_row, column=1).value = 'Feature'
            sheet.cell(row=curr_row, column=1).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=curr_row, column=2).value = signal_type.upper() + ' Baseline'
            sheet.cell(row=curr_row, column=2).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=curr_row, column=3).value = signal_type.upper() + ' Test'
            sheet.cell(row=curr_row, column=3).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=curr_row, column=4).value = signal_type.upper() + ' % Difference'
            sheet.cell(row=curr_row, column=4).font = openpyxl.styles.Font(bold=True)

            feature_types = list(ml_features[signal_type]['baseline_data'].keys())

            for feature in feature_types: 
                curr_row += 1

                baseline_value = categories['baseline_data'].get(feature, 'N/A')
                test_value = categories['test_data'].get(feature, 'N/A')
                percent_diff = categories['percent_difference'].get(feature, 'N/A')

                baseline_value = f"{baseline_value:.5f}" if baseline_value != 'N/A' else 'N/A'
                test_value = f"{test_value:.5f}" if test_value != 'N/A' else 'N/A'
                percent_diff = f"{percent_diff:.5f}%" if percent_diff != 'N/A' else 'N/A%'

                sheet.cell(row=curr_row, column=1).value = feature
                sheet.cell(row=curr_row, column=2).value = baseline_value
                sheet.cell(row=curr_row, column=3).value = test_value
                sheet.cell(row=curr_row, column=4).value = percent_diff

        temp_img = f"shap_waterfall.png"

        fig = ml_prediction['ecg'].get('fig', None)
        if fig is not None:
            fig.savefig(temp_img, dpi=100, bbox_inches='tight')

            img = Image(temp_img)

            ws.add_image(img, f'A{curr_row+3}') 
    
    wb.save(excel_file)

    if os.path.exists(temp_img):
            os.remove(temp_img)


# Saves the stat analysis to excel file
def save_stats_results_to_excel(excel_file, stats_data):
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "Stats Results" not in wb.sheetnames:
        ws = wb.create_sheet(title="Stats Results")
    else:
        ws = wb["Stats Results"]
        ws.delete_rows(1, ws.max_row)
    
    sheet = wb["Stats Results"]
    
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20

    curr_row = 1
    for signal_type, categories in stats_data.items():

        if categories['baseline'] is not None:
            sheet.cell(row=curr_row, column=1).value = 'Stat Type'
            sheet.cell(row=curr_row, column=1).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=curr_row, column=2).value = signal_type.upper() + ' Baseline'
            sheet.cell(row=curr_row, column=2).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=curr_row, column=3).value = signal_type.upper() + ' Test'
            sheet.cell(row=curr_row, column=3).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=curr_row, column=4).value = signal_type.upper() + ' Difference'
            sheet.cell(row=curr_row, column=4).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=curr_row, column=5).value = 'Flag'
            sheet.cell(row=curr_row, column=5).font = openpyxl.styles.Font(bold=True)

            stat_types = list(categories['baseline'].keys())

            for stat in stat_types: 
                curr_row += 1

                baseline_value = categories['baseline'].get(stat, 'N/A')
                test_value = categories['test'].get(stat, 'N/A')
                diff = categories['diff'].get(stat, 'N/A')
                flag = categories['flags'].get(stat, 'N/A')

                baseline_value = f"{baseline_value:.5f}" if baseline_value != 'N/A' else 'N/A'
                test_value = f"{test_value:.5f}" if test_value != 'N/A' else 'N/A'
                diff = f"{diff:.5f}%" if diff != 'N/A' else 'N/A'

                sheet.cell(row=curr_row, column=1).value = stat
                sheet.cell(row=curr_row, column=2).value = baseline_value
                sheet.cell(row=curr_row, column=3).value = test_value
                sheet.cell(row=curr_row, column=4).value = diff
                sheet.cell(row=curr_row, column=5).value = flag
            
            curr_row += 2
                
    
    wb.save(excel_file)
