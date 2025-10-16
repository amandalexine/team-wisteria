import os
from openpyxl import load_workbook, Workbook
import platform
import threading
import numpy as np
import pygame
from test_sequence import set_computer_volume # imported from Jonah's code
import test_sequence
import openpyxl
from openpyxl.styles import Alignment
import datetime

pygame.init()
pygame.mixer.init()

# Global variables
device_list = []
mac_options = []
app = 0
final_volume_db = -30  # Starting volume level
filename = ''
filepath = ''
directory = ''
box=0

    
def print_tb(texts:str, message_box):
    message_box.config(text=texts)

#------------------------------------------------SINE WAVE CREATION--------------------------
def create_sine_wave(freq, db_volume, duration=1):
    linear_volume = db_to_linear(db_volume)
    sample_rate = 44100
    t = np.linspace(0, duration, int(sample_rate * duration), False)
    wave = np.sin(freq * 2 * np.pi * t) * (2**15 - 1) * linear_volume
    wave = wave.astype(np.int16)
    contiguous_wave = np.ascontiguousarray(np.array([wave, wave]).T, dtype=np.int16)
    sound = pygame.sndarray.make_sound(contiguous_wave)
    return sound
 
def db_to_linear(db):
    return 10 ** (db / 20)


def check_utilities():
    print(f"Checking utilities")
    # Check if all utilites are within the folder
    if not os.path.exists("audio_files"):
        os.makedirs('audio_files')

    if not os.path.exists("Patient Records"):
        os.makedirs('Patient Records')

    if not os.path.exists('Utilities'):
        os.makedirs('Utilities')

    if not os.path.exists('Utilities/Bitalino_Devices.txt'):
        create = open('Utilities/Bitalino_Devices.txt', 'x')
        create.close()

    if not os.path.exists('Utilities/ss_logo.ico'):
        return -1
    
    if not os.path.exists('Utilities/sound_sense_logo.png'):
        return -1 
    
    return 0

# ------------------------------------------------SOUND FUNCTIONS------------------------------------------------------------
def play_beep_sound(volume_db):
    beep = create_sine_wave(1000, volume_db, duration=1)
    beep.play()


# ------------------------------------------------BLUETOOTH FUNCTIONS------------------------------------------------------------
def clear_saved_list(mac_combobox, message_box, check):
    global mac_options
 
    if check == 1: # clear everything
        with open('Utilities/Bitalino_Devices.txt', 'w') as file:
            file.write('')
        mac_options = []
        mac_combobox['values'] = []
        print_tb("List cleared", message_box)
    else: # clear combobox only
        mac_options = []
        mac_combobox['values'] = []

def save_to_existing_BITalino_bluetooth_devices(macOption):
    # check if macOption already exsists in text file
    with open('Utilities/Bitalino_Devices.txt', 'r') as file:
        for line in file:
            if line.strip():
                # Split each line by comma and strip whitespace
                parts = line.strip().split(',')
               
                # Remove single quotes from each part
                parts = [part.replace("'", "") for part in parts]
               
                # Create a tuple from the modified parts
                # Assuming there are always two parts
                tuple_from_line = (parts[0].strip(), parts[1].strip())
               
                if tuple_from_line[1] == macOption:
                    return
 
    saved_option = None  # Initialize as None to handle the case if not found
    for i in device_list:
        if i[1] == macOption:
            saved_option = i
            break
   
    if saved_option is not None:
        # Convert saved_option tuple to a string
        saved_option_str = ','.join(map(str, saved_option))
 
        with open('Utilities/Bitalino_Devices.txt', 'a') as file:
            file.write('\n' + saved_option_str)
 
def get_existing_BITalino_bluetooth_devices(mac_combobox):
    global device_list
    file_path = 'Utilities/Bitalino_Devices.txt'

    with open(file_path, 'r') as file:
        for line in file:
            if line.strip():
                # Split each line by comma and strip whitespace
                parts = line.strip().split(',')
            
                # Remove single quotes from each part
                parts = [part.replace("'", "") for part in parts]
            
                # Create a tuple from the modified parts
                # Assuming there are always two parts
                tuple_from_line = (parts[0].strip(), parts[1].strip())
            
                # Append the tuple to the list
                device_list.append(tuple_from_line)

 
    for i in device_list:
        mac_options.append(i[1])
       
    mac_combobox['values'] = mac_options
 
def bluetooth_scan(mac_combobox, message_box, mac_options):
    global device_list
    clear_saved_list(mac_combobox, message_box, 0)
    # The function to be executed by the thread: performs Bluetooth device discovery.
    # Used from: https://github.com/BITalinoWorld/revolution-python-api/blob/master/bitalino.py
    try:
        import bluetooth
    except Exception as e:
        print_tb("Bluetooth failed with error: " + str(e), message_box)
        return
 
    try:
        device_list = bluetooth.discover_devices(lookup_names=True)
        for i in device_list:
            mac_options.append(i[1])
       
        mac_combobox['values'] = mac_options
        print_tb("Scan complete", message_box)
    except Exception as e:
        print_tb("Scanning failed with error: " + str(e), message_box)
 
def find_bluetooth_devices(mac_combobox, message_box):
    if platform.system() not in ["Windows", "Linux"]:
        print_tb("Your platform does not support bluetooth connections", message_box)
        return
 
    print_tb("Scanning now. Please wait...", message_box)
 
    # Start the scan in a separate thread to allow GUI to run simultaneosuly
    scan_thread = threading.Thread(target=bluetooth_scan, args=(mac_combobox, message_box, mac_options))
    scan_thread.start()

# ------------------------------------------------USER INPUT FUNCTIONS-----------------------------------------------------------
def checksubstr(str, substr):
    if substr in str:
        return True
    else:
        return False
    
# Setup patient file
def save_input(name_entry, age_entry, contact_info_entry, label):
    global filename
    global filepath
    global directory
    returning_patient = True

    name = name_entry.get().lower()
    age = age_entry.get()
    contact_info = contact_info_entry.get()

    if name:
        try:
            x = os.getcwd()
            directory = x.replace('\\', '/') 
            filepath = f"Patient Records/Patient_{name}/"

            full_directory = os.path.join(directory, filepath)
            if not os.path.exists(full_directory):
                os.makedirs(full_directory)
                returning_patient = False
                print(f"Created new folder for patient {name}")
            else:
                print(f"Folder already exists for patient {name}")

            patient_files = os.listdir(full_directory)
            existing_numbers = []

            for file in patient_files:
                if file.endswith('.xlsx') and file.startswith(f"{name}_"):
                    try:
                        num = int(file.split('_')[-1].split('.')[0])  # Get the number part from the filename
                        existing_numbers.append(num)
                    except ValueError:
                        continue

            next_number = max(existing_numbers, default=0) + 1
            filename = f"{name}_{next_number}.xlsx"

            filepath_full = os.path.join(full_directory, filename)

            if os.path.exists(filepath_full):
                workbook = load_workbook(filepath_full)
                returning_patient = True
                print(f"Returning patient: {returning_patient}")
            else:
                workbook = Workbook()
                workbook.remove(workbook.active)
                print(f"Returning patient: {returning_patient}")

            sheets = workbook.sheetnames
            if 'Patient Info' not in sheets:
                workbook.create_sheet('Patient Info')
            if 'Recording Info' not in sheets:
                workbook.create_sheet('Recording Info')
            if 'Baseline Data' not in sheets:
                workbook.create_sheet('Baseline Data')
            if 'Test Data' not in sheets:
                workbook.create_sheet('Test Data')
            if 'ML Results' not in sheets:
                workbook.create_sheet('ML Results')
            if 'Signal Graphs' not in sheets:
                workbook.create_sheet('Signal Graphs')
            if 'Stats Results' not in sheets:
                workbook.create_sheet('Stats Results')
            if 'Stats Graphs' not in sheets:
                workbook.create_sheet('Stats Graphs')

            sheet = workbook['Patient Info']
            sheet.cell(row=1, column=1).value = 'Subject #:'
            sheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=1, column=2).value = name
            sheet.cell(row=2, column=1).value = 'Age:'
            sheet.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=2, column=2).value = age
            sheet.cell(row=3, column=1).value = 'Contact Info:'
            sheet.cell(row=3, column=1).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=3, column=2).value = contact_info

            sheet.column_dimensions['A'].width = 18
            sheet.column_dimensions['B'].width = 40
            for row in sheet['B']:
                row.alignment = Alignment(horizontal='left')

            sheet2 = workbook['Test Data']  
            sheet2.cell(row=1, column=1).value = 'EMG'
            sheet2.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
            sheet2.cell(row=1, column=2).value = 'ECG'
            sheet2.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
            sheet2.cell(row=1, column=3).value = 'EDA'
            sheet2.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)

            sheet3 = workbook['Baseline Data']  
            sheet3.cell(row=1, column=1).value = 'EMG'
            sheet3.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
            sheet3.cell(row=1, column=2).value = 'ECG'
            sheet3.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
            sheet3.cell(row=1, column=3).value = 'EDA'
            sheet3.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)

            workbook.save(filepath_full)

            if os.path.exists(filepath_full):
                print(f"Workbook saved successfully at {filepath_full}")
            else:
                print(f"Error: Workbook was not saved at {filepath_full}")

            label.config(text="Information saved successfully!")
        except Exception as e:
            print(f"An error occurred: {str(e)}")
            label.config(text=f"An error occurred: {str(e)}")
    else:
        label.config(text="Please enter all fields!")

# Saving recording info options
def save_recording_info(filepath, filename, recording_info):

    workbook = openpyxl.load_workbook(filepath + filename)
    sheet = workbook['Recording Info']

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    sheet.cell(row=1, column=1).value = 'Timestamp:'
    sheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=1, column=2).value = timestamp
    sheet.cell(row=2, column=1).value = 'Audio Option:'
    sheet.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=2, column=2).value = recording_info["audio_option"]
    sheet.cell(row=3, column=1).value = 'Duration (sec):'
    sheet.cell(row=3, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=3, column=2).value = recording_info["duration"]
    sheet.cell(row=4, column=1).value = 'Sample Rate (Hz):'
    sheet.cell(row=4, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=4, column=2).value = recording_info["sample_rate"]
    sheet.cell(row=5, column=1).value = 'Decibel Increment (dB):'
    sheet.cell(row=5, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=5, column=2).value = recording_info["di_option"]
    sheet.cell(row=6, column=1).value = 'Time Increment (sec):'
    sheet.cell(row=6, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=6, column=2).value = recording_info["time_option"]
    sheet.cell(row=7, column=1).value = 'EMG Recorded:'
    sheet.cell(row=7, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=7, column=2).value = recording_info["signals"][0]
    sheet.cell(row=8, column=1).value = 'ECG Recorded:'
    sheet.cell(row=8, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=8, column=2).value = recording_info["signals"][1]
    sheet.cell(row=9, column=1).value = 'EDA Recorded:'
    sheet.cell(row=9, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=9, column=2).value = recording_info["signals"][2]
    sheet.cell(row=10, column=1).value = 'Device:'
    sheet.cell(row=10, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=10, column=2).value = recording_info["device_option"]

    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 40
    for row in sheet['B']:
        row.alignment = Alignment(horizontal='left')

    workbook.save(filepath + filename)


# ------------------------------------------------BASELINE/TEST SEQUENCE FUNCTIONS-----------------------------------------------

def next_baseline_function(info, controller):
    global device_list
    global app
    global directory
    global filename
    global filepath
    global recording_info

    recording_info = info

    save_recording_info(filepath, filename, recording_info)
    threading.Thread(target=run_baseline_thread, args=(filepath, filename, controller), daemon=True).start()


def run_baseline_thread(filepath, filename, controller):

    # Run different baseline sequence depending on the selected device type
    if recording_info["device_option"] == "ESP32":
        error = test_sequence.run_baseline_sequence_ESP32(filepath, filename, recording_info, controller)
    else:
        error = test_sequence.run_baseline_sequence_bitalino(filepath, filename, recording_info, controller)
    
    controller.after(0, update_after_baseline, error, controller)

def update_after_baseline(error, controller):
    if error == -1:
        controller.show_frame("ErrorPage")
    else:
        controller.show_frame("StartTestPage")

def next_test_sequence_function(controller):
    global device_list
    global app
    global final_volume_db
    global filepath
    global filename
 
    threading.Thread(target=run_test_thread, args=(filepath, filename, controller), daemon=True).start()

def run_test_thread(filepath, filename, controller):
    global recording_info

    # Run different test sequence depending on the selected device type
    if recording_info["device_option"] == "ESP32":
        error = test_sequence.run_test_sequence_ESP32(filepath, filename, final_volume_db, recording_info, controller)
    else:
        error = test_sequence.run_test_sequence_bitalino(filepath, filename, final_volume_db, recording_info, controller)
    
    controller.after(0, update_after_test, error, controller)

def update_after_test(error, controller):
    if error == -1:
        controller.show_frame("ErrorPage")
    else:
        controller.show_frame("ResultsPage")
        return
