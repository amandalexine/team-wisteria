# --------------------------------SAVE FUNCTION-----------------------------------------
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import os

def save_graphs_to_excel(excel_file, graphs):
    # If the file already exists, open it, otherwise create a new Excel workbook
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()

    # Check if the sheet 'Results' already exists, if not, create it
    if "Results" not in wb.sheetnames:
        ws = wb.create_sheet(title="Results")
    else:
        ws = wb["Results"]

    # Find the next available row to append the graphs
    next_row = len(ws['A']) + 1

    count = 0
    # Save each graph to a temporary file and insert it into the Excel sheet
    for i, graph in enumerate(graphs):
        # Save the graph to a temporary image file
        temp_file = f"graph_{i}.png"
        graph.savefig(temp_file)
        
        # Insert the image into the Excel sheet
        img = Image(temp_file)
        ws.add_image(img, f"A{next_row}")

        # Update the row index for the next graph
        next_row += 35
        count += 1

    # Save the Excel workbook
    wb.save(excel_file)

    # Delete temp files
    for i in range(count):
        temp_file = f"graph_{i}.png"
        os.remove(temp_file)
