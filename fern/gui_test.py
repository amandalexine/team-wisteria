import tkinter as tk
from tkinter import ttk
from tkinter import PhotoImage
import ttkbootstrap as tb
import os
from openpyxl import load_workbook,Workbook
import platform
import threading
import numpy as np
import pygame
import customtkinter as ctk
from test_sequence import set_computer_volume # imported from Jonah's code
import test_sequence
import openpyxl
import sys

pygame.init()
pygame.mixer.init()

# Global variables
device_list = []
mac_options = []
app = 0
final_volume_db = -30  # Starting volume level
filename = ''
directory = ''
box=0


#------------------------------------------------SINE WAVE CREATION--------------------------
def create_sine_wave(freq, db_volume, duration=1):
    linear_volume = db_to_linear(db_volume)
    sample_rate = 44100
    t = np.linspace(0, duration, int(sample_rate * duration), False)
    wave = np.sin(freq * 2 * np.pi * t) * (2**15 - 1) * linear_volume
    wave = wave.astype(np.int16)
    contiguous_wave = np.ascontiguousarray(np.array([wave, wave]).T, dtype=np.int16)
    sound = pygame.sndarray.make_sound(contiguous_wave)
    return sound
 
def db_to_linear(db):
    return 10 ** (db / 20)
 
def play_beep_sound(volume_db):
    beep = create_sine_wave(1000, volume_db, duration=1)
    beep.play()
 
def adjust_volume(event):
    global final_volume_db
    if event.keysym == 'Up':
        final_volume_db = min(final_volume_db + 3, 0)  # Cap at 0 dB
    elif event.keysym == 'Down':
        final_volume_db = max(final_volume_db - 3, -60)  # Limit to -60 dB
    play_beep_sound(final_volume_db)
    volume_label.config(text=f"Current Volume: {final_volume_db} dB")
 
# ------------------------------------------------HELPER FUNCTIONS---------------------------------------------------------------
def print_tb(texts:str, message_box):
    message_box.config(text=texts)
 
def clear_frame():
    # If your widgets are directly on the root window `app`
    for widget in app.winfo_children():
        widget.destroy()

def check_utilities():
    # Check if all utilites are within the folder
    if not os.path.exists("audio_files"):
        os.makedirs('audio_files')

    if not os.path.exists("Patient Records"):
        os.makedirs('Patient Records')

    if not os.path.exists('Utilities'):
        os.makedirs('Utilities')

    if not os.path.exists('Utilities/Bitalino_Devices.txt'):
        create = open('Utilities/Bitalino_Devices.txt', 'x')
        create.close()

    if not os.path.exists('Utilities/ss_logo.ico'):
        return -1
    
    if not os.path.exists('Utilities/sound_sense_logo.png'):
        return -1 
    
    return 0

def error_window():
    global frame
    global app
    app = tb.Window(themename="minty")
    app.title("Noise Sensitivity Test")
    app.state('zoomed')
   
 
    frame = ttk.Frame(app)
    frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
 
    screen_width = app.winfo_screenwidth()
    screen_height = app.winfo_screenheight()
    window_width = screen_width * 0.5  
    window_height = screen_height * 0.5
    x_cordinate = int((screen_width / 2) - (window_width / 2))
    y_cordinate = int((screen_height / 2) - (window_height / 2))

    message_box = tk.Message(frame, text="Error in Program! Please try again!", anchor="center", width=400)
    message_box.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    app.mainloop()

# ------------------------------------------------BLUETOOTH FUNCTIONS------------------------------------------------------------
def clear_saved_list(mac_combobox, message_box, check):
    global mac_options
 
    if check == 1: # clear everything
        with open('Utilities/Bitalino_Devices.txt', 'w') as file:
            file.write('')
        mac_options = []
        mac_combobox['values'] = []
        print_tb("List cleared", message_box)
    else: # clear combobox only
        mac_options = []
        mac_combobox['values'] = []

def save_to_existing_BITalino_bluetooth_devices(macOption):
    # check if macOption already exsists in text file
    with open('Utilities/Bitalino_Devices.txt', 'r') as file:
        for line in file:
            if line.strip():
                # Split each line by comma and strip whitespace
                parts = line.strip().split(',')
               
                # Remove single quotes from each part
                parts = [part.replace("'", "") for part in parts]
               
                # Create a tuple from the modified parts
                # Assuming there are always two parts
                tuple_from_line = (parts[0].strip(), parts[1].strip())
               
                if tuple_from_line[1] == macOption:
                    return
 
    saved_option = None  # Initialize as None to handle the case if not found
    for i in device_list:
        if i[1] == macOption:
            saved_option = i
            break
   
    if saved_option is not None:
        # Convert saved_option tuple to a string
        saved_option_str = ','.join(map(str, saved_option))
 
        with open('Utilities/Bitalino_Devices.txt', 'a') as file:
            file.write('\n' + saved_option_str)
 
def get_existing_BITalino_bluetooth_devices(mac_combobox):
    global device_list
    file_path = 'Utilities/Bitalino_Devices.txt'

    with open(file_path, 'r') as file:
        for line in file:
            if line.strip():
                # Split each line by comma and strip whitespace
                parts = line.strip().split(',')
            
                # Remove single quotes from each part
                parts = [part.replace("'", "") for part in parts]
            
                # Create a tuple from the modified parts
                # Assuming there are always two parts
                tuple_from_line = (parts[0].strip(), parts[1].strip())
            
                # Append the tuple to the list
                device_list.append(tuple_from_line)

 
    for i in device_list:
        mac_options.append(i[1])
       
    mac_combobox['values'] = mac_options
 
def bluetooth_scan(mac_combobox, message_box, mac_options):
    global device_list
    clear_saved_list(mac_combobox, message_box, 0)
    # The function to be executed by the thread: performs Bluetooth device discovery.
    # Used from: https://github.com/BITalinoWorld/revolution-python-api/blob/master/bitalino.py
    try:
        import bluetooth
    except Exception as e:
        print_tb("Bluetooth failed with error: " + str(e), message_box)
        return
 
    try:
        device_list = bluetooth.discover_devices(lookup_names=True)
        for i in device_list:
            mac_options.append(i[1])
       
        mac_combobox['values'] = mac_options
        print_tb("Scan complete", message_box)
    except Exception as e:
        print_tb("Scanning failed with error: " + str(e), message_box)
 
def find_bluetooth_devices(mac_combobox, message_box):
    if platform.system() not in ["Windows", "Linux"]:
        print_tb("Your platform does not support bluetooth connections", message_box)
        return
 
    print_tb("Scanning now. Please wait...", message_box)
 
    # Start the scan in a separate thread to allow GUI to run simultaneosuly
    scan_thread = threading.Thread(target=bluetooth_scan, args=(mac_combobox, message_box, mac_options))
    scan_thread.start()
 
# ------------------------------------------------USER INPUT FUNCTIONS-----------------------------------------------------------
def save_input(name_entry, age_entry, contact_info_entry, label):
    name = name_entry.get()
    age = age_entry.get()
    contact_info = contact_info_entry.get()
    global filename
    global directory
 
    if name:
        try:
            x = os.getcwd()
            directory = x.replace('\\', '/')
            filename = f"{name}_{age}.xlsx"
 
            if os.path.exists(directory + '/Patient Records'):
                if os.path.exists('Patient Records/' + filename):
                    workbook = load_workbook(f"Patient Records/{name}_{age}.xlsx")
                else:
                    workbook = Workbook()
                    workbook.remove(workbook.active)  
            else:
                os.makedirs(directory + '/Patient Records')
                workbook = Workbook()
                workbook.remove(workbook.active)
 
           
            sheets = workbook.sheetnames
            if 'Patient Info' not in sheets:
                workbook.create_sheet('Patient Info')
            if 'Results' not in sheets:
                workbook.create_sheet('Results')
            if 'Baseline Data' not in sheets:
                workbook.create_sheet('Baseline Data')
            if 'Test Data' not in sheets:
                workbook.create_sheet('Test Data')
            if 'Cleaned Heart Rate Data' not in sheets:
                workbook.create_sheet('Cleaned Heart Rate Data')
 
            sheet = workbook['Patient Info']
            sheet.append([name, age, contact_info])
            
            sheet1 = workbook['Test Data']
            sheet1.cell(row = 1, column=1).value = 'EMG'
            sheet1.cell(row = 1, column=1).font = openpyxl.styles.Font(bold=True)
            sheet1.cell(row = 1, column=2).value = 'ECG'
            sheet1.cell(row = 1, column=2).font = openpyxl.styles.Font(bold=True)
            sheet1.cell(row = 1, column=3).value = 'EDA'
            sheet1.cell(row = 1, column=3).font = openpyxl.styles.Font(bold=True)

            sheet2 = workbook['Baseline Data']
            sheet2.cell(row = 1, column=1).value = 'EMG'
            sheet2.cell(row = 1, column=1).font = openpyxl.styles.Font(bold=True)
            sheet2.cell(row = 1, column=2).value = 'ECG'
            sheet2.cell(row = 1, column=2).font = openpyxl.styles.Font(bold=True)
            sheet2.cell(row = 1, column=3).value = 'EDA'
            sheet2.cell(row = 1, column=3).font = openpyxl.styles.Font(bold=True)

            workbook.save('Patient Records/' + filename)
            
            label.config(text="Information saved successfully!")
        except Exception as e:
            label.config(text=f"An error occurred: {str(e)}")
    else:
        label.config(text="Please enter all fields!")
 
# ------------------------------------------------BASELINE/TEST SEQUENCE FUNCTIONS-----------------------------------------------
def set_parameters():
    global frame
    clear_frame()
 
    frame = ttk.Frame(app)
    frame.pack(fill=tk.BOTH, padx=7, pady=7)
    app.title("Noise Sensitivity Test")
    app.state('zoomed')
    app.iconbitmap('Utilities/ss_logo.ico')
 
#-----------------------WAV FILE ---------
    directory = "audio_files"
    wav_files=[file for file in os.listdir(directory) if file.endswith('.wav')]
 
    wav_label = ttk.Label(frame, text=f"Select a Wav File (leave empty is none)")
    wav_label.pack(fill=tk.X, padx=5, pady=5)
 
    wav_entry = ttk.Combobox(frame, values=wav_files, state="readonly")
    wav_entry.pack(fill=tk.X, padx=5, pady=5)
 
#-------------FREQUENCY------------
    fq_label = ttk.Label(frame, text="Sound Frequency (in Hz):")
    fq_label.pack(fill=tk.X, padx=5, pady=5)
    # Create a Combobox for sampling rates
    fq_entry = ttk.Entry(master=frame)
    fq_entry.pack(fill=tk.X, padx=5, pady=5)
 
 #-----------Sound increment (in dB)-----------------

    di_label=ttk.Label(frame,text="Sound increment (in dB) (leave empty for default value of 5 dB):")
    di_label.pack(fill=tk.X,padx=5,pady=5)
    di_entry=ttk.Entry(master=frame)
    di_entry.pack(fill=tk.X,padx=5,pady=5)


#-----------Time increment----------------
    ti_label=ttk.Label(frame,text="Time increment (in seconds) (leave empty for default value of 15 seconds):")
    ti_label.pack(fill=tk.X,padx=5,pady=5)
    ti_entry=ttk.Entry(master=frame)
    ti_entry.pack(fill=tk.X,padx=5,pady=5)   
 
    #Sample rate options
    sample_rate_options = [10, 100, 1000]  # Add your desired sample rates
    sr_label = ttk.Label(frame, text="Sampling Rate (in Hz):")
    sr_label.pack(fill=tk.X, padx=5, pady=5)
    # Create a Combobox for sampling rates
    sr_combobox = ttk.Combobox(frame, values=sample_rate_options, state="readonly")
    sr_combobox.pack(fill=tk.X, padx=5, pady=5)
 
 
#_---------------------------------------------------------------------
    #duration_label = ttk.Label(frame, text="Duration (in seconds):")
    #duration_label.pack(fill=tk.X, padx=5, pady=5)
    #duration_entry = ttk.Entry(frame)
    #duration_entry.pack(fill=tk.X, padx=5, pady=5)
#---------------------------------------------------
    duration_options = [15,30,45,60,75,90,105,120]
    duration_label = ttk.Label(frame, text="Duration (in seconds):")
    duration_label.pack(fill=tk.X, padx=5, pady=5)
    duration_entry = ttk.Combobox(frame, values=duration_options, state="readonly")
    duration_entry.pack(fill=tk.X, padx=5, pady=5)
 
 
#------------------------CHECK BOX----------------------------------
    emg = tk.IntVar()
    ecg = tk.IntVar()
    eda = tk.IntVar()

    op1 = ttk.Checkbutton(frame, text="EMG", variable=emg, onvalue=1, offvalue=0)
    op1.pack(fill=tk.X, padx=5,pady=5)
    op2 = ttk.Checkbutton(frame, text="ECG", variable=ecg, onvalue=1, offvalue=0)
    op2.pack(fill=tk.X, padx=5,pady=5)
    op3 = ttk.Checkbutton(frame, text="EDA", variable=eda, onvalue=1, offvalue=0)
    op3.pack(fill=tk.X, padx=5,pady=5)
 
 
 
#-------------------
 
 
    mac_label = ttk.Label(frame, text="Select a BITalino:")
    mac_label.pack(fill=tk.X, padx=5, pady=5)
    mac_combobox = ttk.Combobox(frame, values=mac_options, state='normal')
    mac_combobox.pack(fill=tk.X, padx=5, pady=5)
    get_existing_BITalino_bluetooth_devices(mac_combobox)
    mac_button = ctk.CTkButton(frame, text="Scan for Bluetooth devices", command=lambda:[find_bluetooth_devices(mac_combobox,
                                                                                                             message_box)])
    mac_button.pack(side=tk.TOP, padx=5, pady=(0, 5), anchor='w')
 
    clear_button = ctk.CTkButton(frame, text="Clear current list", command=lambda:clear_saved_list(mac_combobox, message_box,1))
    clear_button.pack(side=tk.TOP, padx=5, pady=(0, 5), anchor='w')  # Adjust padx and pady as needed
 
    status_label = ttk.Label(frame, text="", anchor="center")
    status_label.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
 
 
    def validate_entries(): # function can be used to validate any of the entries above
        try:
            # Convert entires to integers and check if options are valid
            duration = int(duration_entry.get())
            sample_rate = int(sr_combobox.get())
            
            
            # Check if decibel increase value is valid
            if not di_entry.get() == '':
                try:
                    di_option = int(di_entry.get())
                except:
                    message_box.config("Decibel increment entry must be a valid integer")
                    return
            else:
                di_option = 5


            # Check if time options are valid and if are conflicting
            if not ti_entry.get() == '': # if the time interval entry is NOT empty
                try:
                    time_option = int(ti_entry.get())
                except:
                    message_box.config("Time interval entry must be a valid integer")
                    return
            else: # if the time interval entry is empty, default to a value of 15 seconds
                time_option = 15

            if time_option > duration:
                message_box.config("The time interval cannot be larger than the duration")
                return
            

            # Check to see which signals to graph
            signals = []
            if emg.get() == 1:
                signals.append(True)
            else:
                signals.append(False)
            if ecg.get() == 1:
                signals.append(True)
            else:
                signals.append(False)
            if eda.get() == 1:
                signals.append(True)
            else:
                signals.append(False)


            # Check to see which audio options the user selected 
            if not wav_entry.get() == "" and fq_entry.get() == "": # if wave field is populated and frequency field is empty, use .wav file
                audio_option = wav_entry.get()
            elif not fq_entry.get() == "" and wav_entry == "": # if frequency field is populated and wave field is empty, use frequency entry
                try:
                    audio_option = int(fq_entry.get())
                except:
                    message_box.config(text="If you did not enter a .wave file, the frequency must be specified as a valid integer")
                    return
            elif not fq_entry.get() == "" and not wav_entry == "": # if both are populated use frequency entry as default
                try:
                    audio_option = int(fq_entry.get())
                except:
                    message_box.config(text="If you did not enter a .wave file, the frequency must be specified as a valid integer")
                    return
            else:
                message_box.config(text="One or more invalid parameters. Fix entries and try again.")
                return
            

            # Check to see if a BITalino was selected
            if not mac_combobox.get() == '': # if BITalino field is not empty, continue
                mac_option = mac_combobox.get()
            else: # if empty, throw error message
                message_box.config(text="Must choose a BITalino device before continuing")
                return

            save_to_existing_BITalino_bluetooth_devices(mac_option)
            next_baseline_function(sample_rate, duration, mac_option, audio_option, time_option, di_option, signals)
            
        except Exception as e:
            message_box.config(text="One or more invalid parameters. Fix entries and try again.")
           
   
    message_box = tk.Message(frame, text="", anchor="center", width=400)
    message_box.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
 
    start_button = ctk.CTkButton(frame, text="Start Baseline Collection", command=lambda:validate_entries())
    start_button.pack(side=tk.BOTTOM,pady=10,padx=10)
 
def next_baseline_function(sample_rate:int, duration:int, macOption, audio_option, time_option, di_option, signals):
    global device_list
    global app
    global directory
    global filename
 
    clear_frame()
 
    frame = ttk.Frame(app)  # Recreate the frame
    frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
 
    message_box = tk.Message(frame, text="", anchor="center", width=400)
    message_box.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
 
    macAddress = ''
    for i in device_list:
        if i[1] == macOption:
            macAddress = i[0]
 
    # Hide GUI window and pause execution
    app.withdraw()
    
    # Run baseline sequence
    error = test_sequence.run_baseline_sequence(sample_rate, duration, macAddress, filename, signals)
    if error == -1:
        error_window()
    #print("Hello")

    app.deiconify()
    print_tb("Baseline Vital Collection complete.", message_box)
 
    start_button = ctk.CTkButton(frame, text="Start Test Sequence", command=lambda:next_test_sequence_function(sample_rate, duration, macOption,
                                                                                                            audio_option, time_option, di_option,
                                                                                                            signals))
    start_button.pack(side=tk.BOTTOM,pady=10,padx=10)
    return
 
def next_test_sequence_function(sample_rate:int, duration:int, macOption, audio_option, time_option, di_option, signals):
    global device_list
    global app
    global final_volume_db
    global filename
 
    clear_frame()
 
    frame = ttk.Frame(app)  # Recreate the frame
    frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
 
    message_box = tk.Message(frame, text="", anchor="center", width=400)
    message_box.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
 
    macAddress = ''
    for i in device_list:
        if i[1] == macOption:
            macAddress = i[0]
 
    # Hide GUI window and pause execution
    app.withdraw()

    # Run test sequnce
    error = test_sequence.run_test_sequence(sample_rate, duration, macAddress, final_volume_db, audio_option, filename, time_option, di_option, signals)
    if error == -1:
        error_window()

    app.deiconify()
    print_tb("Test Sequence Collection complete.", message_box)
    return

#-------------------------------------------------VOLUME WINDOW------------------------------------------------------------------
def show_volume_adjustment():
    set_computer_volume(50) # From Jonah's code
    clear_frame()
    app.title("Noise Sensitivity Test")
    app.state('zoomed')
    app.iconbitmap('Utilities/ss_logo.ico')
    global volume_label
    ttk.Label(app, text="Use the Up and Down arrow keys to adjust the volume.\nPress 'Next' when finished.").pack(pady=20)
    volume_label = ttk.Label(app, text=f"Current Volume: {final_volume_db} dB")
    volume_label.pack(pady=10)
   
    next_button = ctk.CTkButton(app, text="Next", command=finalize_volume)
    next_button.pack(pady=20)
 
    app.bind('<KeyPress-Up>', adjust_volume)
    app.bind('<KeyPress-Down>', adjust_volume)
 
def finalize_volume():
    clear_frame()
    global final_volume_db
    ttk.Label(app, text=f"Volume set to {final_volume_db} dB. Configuration is now locked.").pack(pady=20)
 
    next_button = ctk.CTkButton(app, text="Next", command=set_parameters)
    next_button.pack(pady=20)
 
 
#----------Wav file scan---------------------
# ------------------------------------------------STARTING/MAIN WINDOW------------------------------------------------------------
def main():
    # Check all Utilites before procceding
    if check_utilities() == -1:
        error_window()
        sys.exit()
        

    global frame
    global app
    app = tb.Window(themename="minty")
    app.title("Noise Sensitivity Test")
    app.state('zoomed')
    app.iconbitmap('Utilities/ss_logo.ico')
   
 
    frame = ttk.Frame(app)
    frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
 
    img_path = "Utilities/sound_sense_logo.png"
    img = PhotoImage(file=img_path, )
 
    img = img.subsample(2, 2)
                         
    img_label = ttk.Label(frame, image=img, anchor="w")
    img_label.image = img
    img_label.pack(side=tk.BOTTOM, anchor="w", padx=10, pady=10)
 
 
    screen_width = app.winfo_screenwidth()
    screen_height = app.winfo_screenheight()
    window_width = screen_width * 0.5  
    window_height = screen_height * 0.5
    x_cordinate = int((screen_width / 2) - (window_width / 2))
    y_cordinate = int((screen_height / 2) - (window_height / 2))
 
 
    style = ttk.Style(app)
   
    frame = ttk.Frame(app)
    frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
 
    name_label = ttk.Label(frame, text="Subject Reference Number *:")
    name_label.pack(fill=tk.X, padx=5, pady=5)
    name_entry = ttk.Entry(frame)
    name_entry.pack(fill=tk.X, padx=5, pady=5)
 
    age_label = ttk.Label(frame, text="Age:")
    age_label.pack(fill=tk.X, padx=5, pady=5)
    age_entry = ttk.Entry(frame)
    age_entry.pack(fill=tk.X, padx=5, pady=5)
 
 
    contact_info_label = ttk.Label(frame, text="Contact Information:")
    contact_info_label.pack(fill=tk.X, padx=5, pady=5)
    contact_info_entry = ttk.Entry(frame)
    contact_info_entry.pack(fill=tk.X, padx=5, pady=5)
 
    requiered_label = ttk.Label(frame, text="*Requiered information")
    requiered_label.pack(fill=tk.X,pady=20,padx=3)
 
    status_label = ttk.Label(frame, text="", anchor="center")
    status_label.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
 
 
    def both():
        if name_entry.get().strip():
            save_input(name_entry,age_entry,contact_info_entry,status_label)
            show_volume_adjustment()
        else:
            status_label.config(text="Subject Reference Number is required to proceed!")
 
    next_button = ctk.CTkButton(frame, text="Next", command=lambda:[both()])
    next_button.pack(side=tk.BOTTOM,pady=10,padx=10)
 
    app.mainloop()

if __name__ == "__main__":
    main()
    
